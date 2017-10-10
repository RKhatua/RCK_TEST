import openpyxl

book = openpyxl.load_workbook('file1.xlsx')
book1 = openpyxl.load_workbook('file2.xlsx')
#sheet = book.active
sheet_list = book.sheetnames
sheet_list1 = book1.sheetnames
if len(sheet_list) != len(sheet_list1):
    print("No of sheets are not equal in both sheets")
    exec(1)
for i in range(len(sheet_list)):
    if sheet_list[i]  not in sheet_list1:
        print("Sheet Name are different in both sheets")

for sheet_name in sheet_list :
    sheet = book.get_sheet_by_name(sheet_name)
    sheet1 = book1.get_sheet_by_name(sheet_name)

    print("\n")
    print(' ' * 15, "Sheet-Name: ", sheet_name, ' ' * 15)
    print("\n")
    rows = sheet.max_row+1
    if sheet.max_row+1 < sheet1.max_row+1 :
        rows = sheet1.max_row+1
    columns =  sheet.max_column+1
    if sheet.max_column+1 < sheet1.max_column+1:
        columns =  sheet1.max_column+1

    for r in range(1, rows):
        for c in range(1, columns):
            cell = sheet.cell(row=r, column=c)
            cell1 = sheet1.cell(row=r, column=c)
            if cell.value != cell1.value :
                print(str(cell.coordinate))
                print("        Old: ", str(cell.value) + '  New: ' + str(cell1.value))
